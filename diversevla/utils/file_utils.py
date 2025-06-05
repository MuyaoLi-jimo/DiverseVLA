"""
# v 2.4 lmdb
"""

import json
import lmdb
import math
import numpy as np 
import pickle
import rich
from tqdm import tqdm
import os
import shutil
import pathlib
import hashlib
import uuid
from collections import defaultdict
from typing import Union,Optional,Any,List,Dict
import zipfile
from datetime import datetime


def generate_uuid():
    return str(uuid.uuid4())

def generate_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
########################################################################
def load_json_file(file_path: Union[str, pathlib.Path], data_type="dict"):
    """
    Load a JSON file from the given path.

    Args:
        file_path (Union[str, pathlib.Path]): Path to the JSON file.
        data_type (str): Expected data type of the JSON content ("dict" or "list").

    Returns:
        dict or list: Loaded JSON content. Returns an empty dictionary or list if the file does not exist.
    """
    if isinstance(file_path, pathlib.Path):
        file_path = str(file_path)  # Convert pathlib.Path to string

    # Initialize an empty object based on the specified data type
    if data_type == "dict":
        json_file = dict()
    elif data_type == "list":
        json_file = list()
    else:
        raise ValueError("Invalid data type. Expected 'dict' or 'list'.")

    # Check if the file exists
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding="utf-8") as f:
                json_file = json.load(f)  # Load JSON content
        except IOError as e:
            rich.print(f"[red]Failed to open file {file_path}: {e}[/red]")
        except json.JSONDecodeError as e:
            rich.print(f"[red]Error parsing JSON file {file_path}: {e}[/red]")
    else:
        rich.print(f"[yellow]File {file_path} does not exist. Returning an empty file...[/yellow]")

    return json_file

def dump_json_file(json_file, file_path: Union[str, pathlib.Path], indent=4, if_print=True, if_backup=True, if_backup_delete=False):
    """
    Save data to a JSON file with optional backup and logging.

    Args:
        json_file (dict or list): The JSON data to save.
        file_path (Union[str, pathlib.Path]): Path to save the JSON file.
        indent (int): Indentation level for formatting the JSON file (default is 4).
        if_print (bool): Whether to print status messages (default is True).
        if_backup (bool): Whether to create a backup before writing (default is True).
        if_backup_delete (bool): Whether to delete the backup after writing (default is False).
    """
    if isinstance(file_path, pathlib.Path):
        file_path = str(file_path)  # Convert pathlib.Path to string

    backup_path = file_path + ".bak"  # Define the backup file path

    # Create a backup if the file exists and backup is enabled
    if os.path.exists(file_path) and if_backup:
        shutil.copy(file_path, backup_path)

    before_exist = os.path.exists(file_path)  # Check if the file existed before writing

    try:
        # Write JSON data to file
        with open(file_path, 'w', encoding="utf-8") as f:
            json.dump(json_file, f, indent=indent, ensure_ascii=False)

        # Print status messages
        if before_exist and if_print:
            rich.print(f"[yellow]Updated {file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]Created {file_path}[/green]")

    except IOError as e:
        # Restore from backup if writing fails
        if os.path.exists(backup_path) and if_backup:
            shutil.copy(backup_path, file_path)
            if if_print:
                rich.print(f"[red]Failed to write {file_path}. Restored from backup: {e}[/red]")
        else:
            if if_print:
                rich.print(f"[red]Failed to write {file_path} and no backup available: {e}[/red]")

    finally:
        # Cleanup: Delete the backup file if required
        if if_backup:
            if os.path.exists(backup_path) and if_backup_delete:
                os.remove(backup_path)
            elif not os.path.exists(backup_path) and not if_backup_delete:  # If the file was initially empty, create a backup
                shutil.copy(file_path, backup_path)

def dump_jsonl(jsonl_file:list,file_path:Union[str , pathlib.Path],if_print=True):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        with open(file_path, 'w',encoding="utf-8") as f:
            for entry in jsonl_file:
                json_str = json.dumps(entry,ensure_ascii=False)
                f.write(json_str + '\n') 
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        print(f"[red]文件{file_path}写入失败，{e}[/red]") 

def split_dump_jsonl(jsonl_file:list,file_path:Union[str , pathlib.Path],split_num = 1, if_print=True):
    # 确保 file_path 是字符串类型
    if isinstance(file_path, pathlib.Path):
        file_path = str(file_path)
    
    # 检查原文件是否存在
    before_exist = os.path.exists(file_path)
    
    # 计算每份数据的大小
    chunk_size = len(jsonl_file) // split_num
    remainder = len(jsonl_file) % split_num  # 计算余数，用来均衡每份的大小

    # 将数据切分成 5 份
    chunks = []
    start_idx = 0
    for i in range(split_num):
        end_idx = start_idx + chunk_size + (1 if i < remainder else 0)  # 如果有余数，前几个切片多分一个
        chunks.append(jsonl_file[start_idx:end_idx])
        start_idx = end_idx
    
    # 写入每份数据到文件
    for i, chunk in enumerate(chunks):
        try:
            # 构造文件路径
            chunk_file_path = file_path[:-6] + f"{i}.jsonl"
            with open(chunk_file_path, 'w', encoding="utf-8") as f:
                for entry in chunk:
                    json_str = json.dumps(entry, ensure_ascii=False)
                    f.write(json_str + '\n')

            # 打印文件创建或更新信息
            if before_exist and if_print:
                rich.print(f"[yellow]更新{chunk_file_path}[/yellow]")
            elif if_print:
                rich.print(f"[green]创建{chunk_file_path}[/green]")
        except IOError as e:
            print(f"[red]文件{chunk_file_path}写入失败，{e}[/red]")

def load_jsonl(file_path:Union[str , pathlib.Path]):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    jsonl_file = []
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r') as f:
                for line in f:
                    jsonl_file.append(json.loads(line))
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}")
        except json.JSONDecodeError as e:
            rich.print(f"[red]解析 JSON 文件时出错：{e}")
    else:
        rich.print(f"[yellow]{file_path}文件不存在，正在传入空文件...[/yellow]")
    return jsonl_file 
                
class JsonlProcessor:
    def __init__(self, file_path:Union[str , pathlib.Path],
                 if_backup = True,
                 if_print=True
                 ):
        
        self.file_path = file_path if not isinstance(file_path,pathlib.Path) else str(file_path)
        
        self.if_print = if_print
        self.if_backup = if_backup

        self._mode = ""

        self._read_file = None
        self._write_file = None
        self._read_position = 0
        self.lines = 0

    @property
    def bak_file_path(self):
        return str(self.file_path) + ".bak"
    
    def exists(self):
        return os.path.exists(self.file_path)

    def len(self):
        file_length = 0
        if not self.exists():
            return file_length
        if self.lines == 0:
            with open(self.file_path, 'r', encoding='utf-8') as file:
                while file.readline():
                    file_length+=1
            self.lines = file_length
        return self.lines

    def close(self,mode = "rw"):
        # 关闭文件资源
        if "r" in mode:
            if self._write_file:
                self._write_file.close()
                self._write_file = None
        if "w" in mode:
            if self._read_file:
                self._read_file.close()
                self._read_file = None
            self.lines = 0
        

    def reset(self, file_path:Union[str , pathlib.Path]):
        self.close()
        self.file_path = file_path if not isinstance(file_path,pathlib.Path) else str(file_path)


    def load_line(self,fast:bool=False):
        if not fast:
            if not self.exists():
                rich.print(f"[yellow]{self.file_path}文件不存在,返回{None}")
                return None
            if self._mode != "r":
                self.close("r")
                
        if not self._read_file:
            self._read_file = open(self.file_path, 'r', encoding='utf-8')
            
        if not fast:
            self._read_file.seek(self._read_position)
            self._mode = "r"
       
        try:
            line = self._read_file.readline()
            self._read_position = self._read_file.tell()
            if not line:
                self.close()
                return None
            return json.loads(line.strip())
        except json.JSONDecodeError as e:
            self.close()
            rich.print(f"[red]文件{self.file_path}解析出现错误：{e}")
            return None
        except IOError as e:
            self.close()
            rich.print(f"[red]无法打开文件{self.file_path}：{e}")
            return None
    
    def load_lines(self):
        """获取jsonl中的line，直到结尾"""
        lines = []
        while True:
            line = self.load_line()
            if line ==None:
                break
            lines.append(line)
        return lines
        

    def load_restart(self):
        self.close(mode="r")
        self._read_position = 0
         
    def dump_line(self, data,fast:bool=False):
        if not isinstance(data,dict) and not isinstance(data,list):
            raise ValueError("数据类型不对")
        if not fast:
            # 备份
            if self.len() % 50 == 1 and self.if_backup:
                shutil.copy(self.file_path, self.bak_file_path)
            self._mode = "a"
            # 如果模型尚未打开
        if not self._write_file:
            self._write_file = open(self.file_path, 'a', encoding='utf-8')
        try:
            json_line = json.dumps(data,ensure_ascii=False)
            self._write_file.write(json_line + '\n')
            self._write_file.flush()
            self.lines += 1  
            return True
        except Exception as e:
            
            if os.path.exists(self.bak_file_path) and self.if_backup:
                shutil.copy(self.bak_file_path, self.file_path)
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，已从备份恢复原文件: {e}[/red]")
            else:
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，且无备份可用：{e}[/red]") 
            return False

    def dump_lines(self,datas):
        if not isinstance(datas,list):
            raise ValueError("数据类型不对")
        if self.if_backup and os.path.exists(self.file_path):
            shutil.copy(self.file_path, self.bak_file_path)
        self._mode = "a"
        if not self._write_file:
            self._write_file = open(self.file_path, 'a', encoding='utf-8')
        try:
            self.len()
            for data in datas:
                json_line = json.dumps(data,ensure_ascii=False)
                self._write_file.write(json_line + '\n')
                self.lines += 1  
            self._write_file.flush()
            return True
        except Exception as e:
            if os.path.exists(self.bak_file_path) and self.if_backup:
                shutil.copy(self.bak_file_path, self.file_path)
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，已从备份恢复原文件: {e}[/red]")
            else:
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，且无备份可用：{e}[/red]") 
                return False
            
    def dump_restart(self):
        self.close()
        self._mode= "w"
        with open(self.file_path, 'w', encoding='utf-8') as file:
            pass 
          
    def load(self):
        jsonl_file = []
        if self.exists():
            try:
                with open(self.file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        jsonl_file.append(json.loads(line))
            except IOError as e:
                rich.print(f"[red]无法打开文件：{e}")
            except json.JSONDecodeError as e:
                rich.print(f"[red]解析 JSON 文件时出错：{e}")
        else:
            rich.print(f"[yellow]{self.file_path}文件不存在，正在传入空文件...[/yellow]")
        return jsonl_file

    def dump(self,jsonl_file:list):
        before_exist = self.exists()
        if self.if_backup and before_exist:
            shutil.copy(self.file_path, self.bak_file_path)
        try:
            self.close()
            self._mode = "w"
            with open(self.file_path, 'w', encoding='utf-8') as f:
                for entry in jsonl_file:
                    json_str = json.dumps(entry,ensure_ascii=False)
                    f.write(json_str + '\n') 
                    self.lines += 1
            if before_exist and self.if_print:
                rich.print(f"[yellow]更新{self.file_path}[/yellow]")
            elif self.if_print:
                rich.print(f"[green]创建{self.file_path}[/green]")
            return True
        except Exception as e:
            if os.path.exists(self.bak_file_path) and self.if_backup:
                shutil.copy(self.bak_file_path, self.file_path)
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，已从备份恢复原文件: {e}[/red]")
            else:
                if self.if_print:
                    rich.print(f"[red]文件{self.file_path}写入失败，且无备份可用：{e}[/red]") 
            return False  

def load_npy_file(file_path:Union[str , pathlib.Path]):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    npy_array = np.empty((0,))
    if os.path.exists(file_path):
        try:
            npy_array = np.load(file_path)
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}[/red]")
    else:
         rich.print(f"[yellow]{file_path}文件不存在，传入np.empty((0,))[/yellow]")

    return npy_array

def dump_npy_file(npy_array:np.ndarray, file_path:Union[str , pathlib.Path],if_print = True):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        np.save(file_path,npy_array)
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")

def load_pickle_file(file_path:Union[str , pathlib.Path]):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    pkl_file = {}
    if os.path.exists(file_path):
        try:
            with open(file_path, 'rb') as file:
                # 使用pickle.load加载并反序列化数据
                pkl_file = pickle.load(file)
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}[/red]")
    else:
         rich.print(f"[yellow]{file_path}文件不存在，传入空文件[/yellow]")

    return pkl_file

def dump_pickle_file(pkl_file, file_path:Union[str , pathlib.Path],if_print = True):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        with open(file_path, 'wb') as file:
            # 使用pickle.dump将数据序列化到文件
            pickle.dump(pkl_file, file)
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")

def load_txt_file(file_path:Union[str , pathlib.Path]):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r') as f:
                txt_file = f.read()
        except IOError as e:
            rich.print(f"[red]无法打开文件：{e}[/red]")
    else:
         rich.print(f"[yellow]{file_path}文件不存在，传入空文件[/yellow]")

    return txt_file

def dump_txt_file(file,file_path:Union[str , pathlib.Path],if_print = True):
    if isinstance(file_path,pathlib.Path):
        file_path = str(file_path)
    before_exist = os.path.exists(file_path)
    try:
        with open(file_path, 'w') as f:
            # 使用pickle.dump将数据序列化到文件
            f.write(str(file))
        if before_exist and if_print:
            rich.print(f"[yellow]更新{file_path}[/yellow]")
        elif if_print:
            rich.print(f"[green]创建{file_path}[/green]")
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")

def load_excel_file_to_dict(file_path:Union[str , pathlib.Path],if_print = True):
    """存储成如下格式：
    {
        "sheet_name1":[
            {
                "column1":"",
                "column2":"",
                "column3":"",
    }]}
    """
    import openpyxl
    if isinstance(file_path,str):
        file_path = pathlib.Path(file_path)
    assert file_path.suffix == ".xlsx"
    wb = openpyxl.load_workbook(file_path)
    data = {}
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        sheet_data = []

        for row in rows[1:]:
            row_data = {headers[i]: row[i] for i in range(len(headers))}
            sheet_data.append(row_data)

        data[sheet] = sheet_data
    return data

def dump_excel_file(file:dict, file_path:Union[str , pathlib.Path],if_print = True):
    """转换各种模式为xlsx(excel模式)"""
    import openpyxl
    if isinstance(file_path,str):
        file_path = pathlib.Path(file_path)
    assert file_path.suffix == ".xlsx"
    
    wb = openpyxl.Workbook()
    
    if isinstance(file, dict):
        """
        如果是dict，暂时要求符合如下格式：
        {
            "sheet_name1":[
                {
                    "column1":"",
                    "column2":"",
                    "column3":"",
        }]}
        """

        sheet0 = list(file.values())[0]
        assert isinstance(sheet0, list)
        row0 = sheet0[0]
        assert isinstance(row0,dict)
        item0 = list(row0.values())[0]
        assert isinstance(item0,str)
        # 然后转成DataFrame模式
        wb.remove(wb.active)  # 移除默认创建的空白工作表
        # 遍历 JSON 数据中的每个工作表
        for sheet_name, rows in file.items():
            ws = wb.create_sheet(title=sheet_name)  # 创建新的工作表
            headers = rows[0].keys()  # 假设所有行的键相同，作为表头
            ws.append(list(headers))  # 添加表头
            for row in rows:
                ws.append(list(row.values()))  # 添加数据行
    try:
        wb.save(file_path)
    except IOError as e:
        rich.print(f"[red]文件写入失败：{e}[/red]")
        
    if file_path.exists() and if_print:
        rich.print(f"[yellow]更新{file_path}[/yellow]")
    elif if_print:
        rich.print(f"[green]创建{file_path}[/green]")

##############################################
    
def zip_fold(source_path:Union[str , pathlib.Path], zip_path:Union[str , pathlib.Path]):
    if isinstance(source_path,str):
        source_path = pathlib.Path(source_path)
    if isinstance(zip_path,str):
        zip_path = pathlib.Path(zip_path)
    if not zip_path.exists():
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(source_path):
                for file in files:
                    # 创建ZIP文件中的文件路径，包括其在文件夹中的相对路径
                    zipf.write(os.path.join(root, file),
                            os.path.relpath(os.path.join(root, file), 
                                            os.path.join(source_path, '..')))
        print(f"[red]{zip_path}已经创建")

def unzip_fold(zip_path:Union[str , pathlib.Path],target_fold:Union[str , pathlib.Path]=None):
    if isinstance(zip_path,str):
        zip_path = pathlib.Path(zip_path)
    if type(target_fold) == type(None):
        parent_path = zip_path.parent
        file_name = zip_path.stem
        target_fold = parent_path / file_name
        pass
    elif isinstance(target_fold,str):   
        target_fold = pathlib.Path(target_fold)
    
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(target_fold)

    print(f"[red]{zip_path}解压到{target_fold}")

def rm_folder(folder_path:Union[str , pathlib.Path]):
    if isinstance(folder_path,str):
        folder_path = pathlib.Path(folder_path)
    if folder_path.exists() and folder_path.is_dir():
        shutil.rmtree(folder_path)
        print(f"Folder '{folder_path}' and its contents have been deleted.")
    else:
        print(f"Folder '{folder_path}' does not exist or is not a directory.")

################################################

class LmdbProcessor:
    """
    A class to handle LMDB database operations with support for serialization.
    """
    def __init__(
        self, 
        path: Union[str, pathlib.Path], 
        map_size: int = 1 << 30,  # Default size is 1GB
        readonly: bool = True, 
        lock: bool = True, 
        readahead: bool = False,
        serializer: str = 'pickle'
    ):
        """
        Initializes the LMDB environment.

        Parameters:
            path (Union[str, pathlib.Path]): The path to the LMDB database.
            map_size (int): The maximum size of the database map.
            readonly (bool): Set to True for read-only mode.
            lock (bool): Set to True to lock the database during operations.
            readahead (bool): Set to False to disable readahead (can improve performance in some environments).
            serializer (str): The serialization method used; currently only supports 'pickle'.
        """
        self.path = str(path)
        self.map_size = map_size
        self.readonly = readonly
        self.lock = lock
        self.readahead = readahead
        self.env = lmdb.open(
            self.path, 
            map_size=self.map_size, 
            readonly=self.readonly, 
            lock=self.lock, 
            readahead=self.readahead,
            max_dbs=1,
            create=True  # By default, create=True so if it doesn't exist, it will be created
        )
        self.serializer = serializer

        # Check if we can or need to expand the map size
        # We can only expand if we're not in read-only mode
        if not self.readonly:
            info = self.env.info()
            current_map_size = info["map_size"]
            # If the requested map_size is larger than the current one, expand
            if map_size > current_map_size:
                self.env.set_mapsize(map_size)

    def serialize(self, value: Any) -> bytes:
        if self.serializer == 'pickle':
            return pickle.dumps(value)
        raise ValueError(f"Unsupported serializer: {self.serializer}")

    def deserialize(self, data: bytes) -> Any:
        if self.serializer == 'pickle':
            return pickle.loads(data)
        raise ValueError(f"Unsupported serializer: {self.serializer}")

    def insert(self, key: str, value: Any) -> bool:
        if self.readonly:
            raise PermissionError("Cannot insert in readonly mode.")
        key_encoded = key.encode('utf-8')
        value_encoded = self.serialize(value)
        with self.env.begin(write=True) as txn:
            return txn.put(key_encoded, value_encoded)

    def insert_bulk(self, kv_dict: Dict[str, Any]) -> None:
        if self.readonly:
            raise PermissionError("Cannot insert in readonly mode.")
        with self.env.begin(write=True) as txn:
            for key, value in kv_dict.items():
                key_encoded = key.encode('utf-8')
                value_encoded = self.serialize(value)
                txn.put(key_encoded, value_encoded)

    def delete(self, key: str) -> bool:
        if self.readonly:
            raise PermissionError("Cannot delete in readonly mode.")
        key_encoded = key.encode('utf-8')
        with self.env.begin(write=True) as txn:
            return txn.delete(key_encoded)

    def get(self, key: str) -> Optional[Any]:
        key_encoded = key.encode('utf-8')
        with self.env.begin() as txn:
            data = txn.get(key_encoded)
        if data is not None:
            return self.deserialize(data)
        return None

    def get_all_keys(self) -> List[str]:
        with self.env.begin() as txn:
            with txn.cursor() as cursor:
                return [key.decode('utf-8') for key, _ in cursor]

    def get_info(self) -> dict:
        with self.env.begin() as txn:
            with txn.cursor() as cursor:
                return {
                    key.decode('utf-8'): self.deserialize(value)
                    for key, value in cursor
                }

    def close(self):
        self.env.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.close()


class ShardedLmdbManager:
    """
    Manages multiple LMDB shards to handle very large datasets.
    It provides an interface similar to a single LMDB, but internally
    distributes data across multiple shards.
    """

    def __init__(
        self,
        base_path: Union[str, pathlib.Path],
        num_shards: int = None,
        map_size: int = None,
        map_size_per_shard: int = 1 << 35,  # Default 32GB per shard
        readonly: bool = False,
        lock: bool = True,
        readahead: bool = False,
        serializer: str = 'pickle',
    ):
        """
        Initializes multiple LmdbProcessor instances (shards). If shard directories
        already exist, we re-open them. Otherwise, we create them from scratch.

        Parameters:
            base_path (str|Path): The directory in which multiple LMDB shards will be located.
            num_shards (int|None): Number of shards. If shards already exist on disk,
                                   we can infer the count automatically.
            map_size (int|None): An overall map size to distribute among shards.
                                 If provided, num_shards = map_size // map_size_per_shard
            map_size_per_shard (int): The map_size for each shard (default: 5GB).
            readonly (bool): Whether shards are opened in read-only mode.
            lock (bool): Whether to lock the database during operations.
            readahead (bool): Whether to enable or disable readahead on the database.
            serializer (str): Serialization method for each shard ('pickle' by default).
        """
        self.base_path = str(base_path)
        os.makedirs(self.base_path, exist_ok=True)
        
        # If there's an existing set of shards on disk, detect them:
        existing_shards = [
            d for d in os.listdir(self.base_path)
            if d.startswith("shard_") and os.path.isdir(os.path.join(self.base_path, d))
        ]
        existing_shards = sorted(existing_shards, key=lambda x: int(x.split("_")[-1]))
        num_existing_shards = len(existing_shards)

        # Figure out final num_shards:
        if map_size is not None:
            # If a total map_size is given, compute how many shards that implies:
            computed = math.ceil(map_size / map_size_per_shard)
            # If the user also gave num_shards explicitly, they must match
            if num_shards is not None and num_shards != computed:
                raise ValueError(
                    f"Mismatch: map_size // map_size_per_shard = {computed}, "
                    f"but num_shards={num_shards} was provided."
                )
            num_shards = computed
        elif num_shards is None:
            # If the user didn't specify num_shards AND didn't specify map_size,
            # but some shards exist, assume we just want to open those.
            if num_existing_shards > 0:
                num_shards = num_existing_shards
            else:
                raise ValueError(
                    "Cannot determine num_shards automatically because no shards exist on disk "
                    "and map_size was not provided."
                )

        self.num_shards = num_shards
        self.shards: List[LmdbProcessor] = []

        # If there are existing shards on disk:
        if num_existing_shards > 0:
            # If existing count doesn't match what we want, raise or handle it
            if num_existing_shards != self.num_shards:
                raise ValueError(
                    f"Found {num_existing_shards} existing shards on disk, but "
                    f"num_shards={self.num_shards} was requested."
                )
            # Re-open existing shards
            for i in range(self.num_shards):
                shard_path = os.path.join(self.base_path, f"shard_{i}")
                processor = LmdbProcessor(
                    path=shard_path,
                    map_size=map_size_per_shard,
                    readonly=readonly,
                    lock=lock,
                    readahead=readahead,
                    serializer=serializer
                )
                self.shards.append(processor)
        else:
            # No existing shards, so create them
            for i in range(self.num_shards):
                shard_path = os.path.join(self.base_path, f"shard_{i}")
                os.makedirs(shard_path, exist_ok=True)
                processor = LmdbProcessor(
                    path=shard_path,
                    map_size=map_size_per_shard,
                    readonly=readonly,
                    lock=lock,
                    readahead=readahead,
                    serializer=serializer
                )
                self.shards.append(processor)

    def _shard_index(self, key: str) -> int:
        """
        Calculates the shard index for the given key, for example by:
          - Taking the MD5 hash of the key
          - Converting it to an integer
          - Using modulo to distribute among shards
        """
        h = hashlib.md5(key.encode('utf-8')).hexdigest()
        return int(h, 16) % self.num_shards

    def insert(self, key: str, value: Any) -> bool:
        """
        Inserts a single key-value pair into the appropriate shard.

        Raises:
            PermissionError: If the manager was initialized in read-only mode.

        Returns:
            bool: True if the key was successfully inserted/updated.
        """
        if self.shards and self.shards[0].readonly:
            raise PermissionError("Cannot insert in read-only mode (ShardedLmdbManager).")

        shard_idx = self._shard_index(key)
        return self.shards[shard_idx].insert(key, value)

    def insert_bulk(self, kv_dict: Dict[str, Any]) -> None:
        """
        Inserts multiple key-value pairs into the appropriate shards, each in a single transaction per shard.

        Raises:
            PermissionError: If the manager was initialized in read-only mode.
        """
        if self.shards and self.shards[0].readonly:
            raise PermissionError("Cannot insert in read-only mode (ShardedLmdbManager).")

        shard_batches = defaultdict(dict)
        for key, value in kv_dict.items():
            idx = self._shard_index(key)
            shard_batches[idx][key] = value

        for idx, shard_dict in shard_batches.items():
            self.shards[idx].insert_bulk(shard_dict)

    def get(self, key: str) -> Optional[Any]:
        """
        Retrieves the value corresponding to the given key from the appropriate shard.

        Returns:
            The deserialized value if the key exists, or None otherwise.
        """
        shard_idx = self._shard_index(key)
        return self.shards[shard_idx].get(key)

    def delete(self, key: str) -> bool:
        """
        Deletes a key-value pair from the appropriate shard.

        Raises:
            PermissionError: If the manager was initialized in read-only mode.

        Returns:
            bool: True if the key existed and was deleted, False otherwise.
        """
        if self.shards and self.shards[0].readonly:
            raise PermissionError("Cannot delete in read-only mode (ShardedLmdbManager).")

        shard_idx = self._shard_index(key)
        return self.shards[shard_idx].delete(key)

    def get_all_keys(self) -> List[str]:
        """
        Retrieves all keys from all shards and returns them as a combined list.
        Duplicates (if any) are removed.

        Returns:
            A list of unique key strings across all shards.
        """
        all_keys = []
        for processor in self.shards:
            all_keys.extend(processor.get_all_keys())
        return list(set(all_keys))

    def close(self):
        """
        Closes all shard environments.
        """
        for processor in self.shards:
            processor.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.close()

    def merge_all_shards_to_single(self, new_path: Union[str, pathlib.Path], new_map_size: int = 1 << 32):
        """
        Merges all data from these shards into a single LMDB environment.

        Parameters:
            new_path (str|Path): The target path for the merged LMDB environment.
            new_map_size (int): The map size for the new environment (default: 4GB).

        This method creates a new (writable) LmdbProcessor at 'new_path'
        and copies all key-value pairs from each shard into the new environment.
        """
        new_proc = LmdbProcessor(
            path=new_path,
            map_size=new_map_size,
            readonly=False
        )

        for shard_proc in self.shards:
            shard_data = shard_proc.get_info()
            for k, v in shard_data.items():
                new_proc.insert(k, v)

        new_proc.close()
        print(f"Data merged successfully. The new LMDB environment is located at: {new_path}")
################################################

if __name__ == "__main__":
    print(generate_uuid())
    exit()
    jp = JsonlProcessor("temp/1.jsonl")
    jp.dump_restart()
    list1 = [a for a in range(10)]
    jp.dump_lines(list1)
    lines = jp.load_lines()
    print(lines)
    list2 = [a for a in range(10,20)]
    jp.dump_lines(list2)
    lines = jp.load_lines()
    print(lines)
    jp.dump_lines(list2)
    lines = jp.load_lines()
    print(lines)
    jp.close()